#! /usr/bin/env python3
# coding: utf-8

import openpyxl as xl
import pandas as pd


def boundaries_range(first_cell_column,
                     first_cell_row,
                     last_cell_column,
                     last_cell_row):
    """Convert 2 cell coordinates into a range as a string"""
    
    # Input checks
    if (first_cell_column > last_cell_column or
        last_cell_row > last_cell_row):
        raise ValueError("Invalid input : wrong cell order for a range")
    
    # Conversion process
    letter = xl.utils.cell.get_column_letter
    str_first_cell = letter(first_cell_column) + str(first_cell_row)
    str_last_cell = letter(last_cell_column) + str(last_cell_row)
    return str_first_cell + ":" + str_last_cell


def worksheet_clean_range(sheet, cell_range):
    """
    Return a range corresponding to the input range cleaned
    of any fully empty row/column around the data. Empty
    row/column in the middle of the data are not detected.

    Parameters
    ----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        Excel-type worksheet defined as part of a workbook
        in openpyxl.
    cell_range : string
        Excel type range as a string. For example "A1:B2"

    Returns
    -------
    Range as a string, or None if all cells in the range are
    empty or if the range is out of the worksheet bounds.

    """
    
    # Get range boundaries coordinates
    (input_first_col_i,
     input_first_row_i,
     input_last_col_i,
     input_last_row_i) = xl.utils.cell.range_boundaries(cell_range)
    
    # Check if worksheet bounds are not crossed
    # Otherwise, index omissions can happen inside iter_rows and
    # lead to wrong row index
    input_first_col_i = max(input_first_col_i, sheet.min_column)
    input_first_row_i = max(input_first_row_i, sheet.min_row)
    input_last_col_i = min(input_last_col_i, sheet.max_column)
    input_last_row_i = min(input_last_row_i, sheet.max_row)
    if (input_first_col_i > input_last_col_i or 
        input_first_row_i > input_last_row_i):
        raise ValueError("Input range out of worksheet bounds")

    # Get first not-empty row
    cleaned_first_row_i = input_first_row_i
    row_iterable = sheet.iter_rows(min_row=input_first_row_i,
                                   max_row=input_last_row_i,
                                   min_col=input_first_col_i,
                                   max_col=input_last_col_i)
    for i, row in enumerate(row_iterable):
        row_i = input_first_row_i + i
        if all([cell.value is None for cell in row]):
            cleaned_first_row_i = row_i + 1
        else:
            break
        
    # Empty range case
    if cleaned_first_row_i > input_last_row_i:
        return None
    
    # Get last not-empty row
    cleaned_last_row_i = input_last_row_i
    row_iterable = sheet.iter_rows(min_row=cleaned_first_row_i,
                                   max_row=input_last_row_i,
                                   min_col=input_first_col_i,
                                   max_col=input_last_col_i)
    for rev_i, row in enumerate(reversed(list(row_iterable))):
        row_i = input_last_row_i - rev_i
        if all([cell.value is None for cell in row]):
            cleaned_last_row_i = row_i - 1
        else:
            break    
        
    # Get first and last not-empty columns
    cleaned_first_col_i = input_last_col_i # init : only last col filled
    cleaned_last_col_i = input_first_col_i # init : only first col filled
    for row in sheet.iter_rows(min_row=cleaned_first_row_i,
                               max_row=cleaned_last_row_i,
                               min_col=input_first_col_i,
                               max_col=input_last_col_i):
        row_cells_are_empty = [cell.value == None for cell in row]
        try:
            first_nonempty_i = row_cells_are_empty.index(False)
            row_first_col_i = input_first_col_i + first_nonempty_i
            last_nonempty_i = row_cells_are_empty[::-1].index(False)
            row_last_col_i = input_last_col_i - last_nonempty_i
        except:
            pass # no action if an empty row exists in the data
        cleaned_first_col_i = min(cleaned_first_col_i, row_first_col_i)
        cleaned_last_col_i = max(cleaned_last_col_i, row_last_col_i)
        
    # Range construction and return
    return boundaries_range(cleaned_first_col_i,
                            cleaned_first_row_i,
                            cleaned_last_col_i,
                            cleaned_last_row_i)
    

def worksheet_to_dataframe(ws):
    """
    Store the data contained in an Excel worksheet in a pandas.DataFrame
    The worksheet must respect a given format.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to process (data extraction). The sheet must
        respect the following constraints :
            - rectangular data range with :
                - data name as first column
                - data unit as second column
                - data as columns, one column per point
            - otionnal comments before the data, as a group
              of lines without empty row. Cells content will
              be concatenated (whitespace as separator). These
              comments must be separated from the data by at
              least one empty row.

    Returns
    -------
    dataframe : pandas.DataFrame
        Dataframe containing the data of the input worksheet.
        Indexes (points) are defined as tupples of strings :
        (worksheet_name, point_index_starting_at_1)
        Column names are also tupples of strings :
        (data_name, data_unit)
    """
        
    # Clean worksheet range to remove empty rows/column around the data
    all_range = worksheet_clean_range(ws, ws.calculate_dimension())

    # Get range boundaries coordinates
    (first_col_i,
     first_row_i,
     last_col_i,
     last_row_i) = xl.utils.cell.range_boundaries(all_range)

    # Iteration on rows from last to first to find last cell before
    # empty row(s) to identify potential comments
    last_h_row_i = None
    first_d_row_i = first_row_i
    for i, row in enumerate(ws.iter_rows(min_row=first_row_i,
                                            max_row=last_row_i)):
        # comment presence detected and last comment line index saved
        if all([cell.value is None for cell in row]):
            if last_h_row_i == None:    
                last_h_row_i = first_row_i + i - 1
        # Else, if comment previously detected, first non-empty
        # line after comment saved as first data line
        elif last_h_row_i != None:
            first_d_row_i = first_row_i + i
            break
        
    # comments stored as a string
    comments = ""
    for row in ws.iter_rows(min_row=first_row_i,
                               max_row=last_h_row_i):
        line = [str(cell.value) for cell in row if cell.value != None]
        if bool(line) and bool(comments): # Empty list/str == False
            comments += "\n"
        separator = " "
        comments += separator.join(line)
        
    # Cleaning of the data range
    data_range = boundaries_range(first_col_i,
                                  first_d_row_i,
                                  last_col_i,
                                  last_row_i)
    data_range = worksheet_clean_range(ws, data_range)
    (first_d_col_i,
     first_d_row_i,
     last_d_col_i,
     last_d_row_i) = xl.utils.cell.range_boundaries(data_range)
    
    # Data range split to get libs, units & values
    libs_range = boundaries_range(first_d_col_i,
                                  first_d_row_i,
                                  first_d_col_i,
                                  last_d_row_i)
    units_range = boundaries_range(first_d_col_i+1,
                                   first_d_row_i,
                                   first_d_col_i+1,
                                   last_d_row_i)
    vals_range = boundaries_range(first_d_col_i+2,
                                   first_d_row_i,
                                   last_d_col_i,
                                   last_d_row_i)
    libs = [row[0].value for row in ws[libs_range]]
    units = [row[0].value for row in ws[units_range]]
    vals = [[cell.value for cell in row] for row in ws[vals_range]]
    
    # Transposing values to get single data type per column for faster
    # pandas processing
    vals = [*zip(*vals)]
    
    # Defining indexes and column names as tuples, and converting them
    # to multi-indexes
    index_tuples = [(ws.title, str(pt)) for pt in range(1,len(vals)+1)]
    multi_index = pd.MultiIndex.from_tuples(index_tuples)
    column_tuples = [*zip(libs, units)]
    multi_column = pd.MultiIndex.from_tuples(column_tuples)
    
    # Pandas Dataframe creation and return
    return pd.DataFrame(data=vals,
                        index=multi_index,
                        columns=multi_column)


def workbook_to_dataframe(filepath):
    """
    Load a workbook and store the data contained in each sheet
    in a separate pandas.Dataframe. The workbook must respect
    a correct format.

    Parameters
    ----------
    filepath : string
        Filepath to an Excel workbook to load. All worksheets
        are loaded and values are extracted. All sheets must
        respect the following constraints :
            - rectangular data range with :
                - data name as first column
                - data unit as second column
                - data as columns, one column per point
            - otionnal comments before the data, as a group
              of lines without empty row. Cells content will
              be concatenated (whitespace as separator). These
              comments must be separated from the data by at
              least one empty row.

    Returns
    -------
    concat_dataframe : pandas.DataFrame
        Dataframe containing the data of the input workbook.
        Indexes (points) are defined as tupples of strings :
        (worksheet_name, point_index_starting_at_1)
        Column names are also tupples of strings :
        (data_name, data_unit)
    """
    # Workbook loading and closure
    wb = xl.load_workbook(filepath, read_only=True, data_only=True)
    
    # Empty list to store dataframes for each processed worksheet
    sheet_dataframes = []
    
    # Each sheet is processed independently
    for ws in wb.worksheets:
        sheet_dataframes.append(worksheet_to_dataframe(ws))
        
    # Concatenation of the dataframes for the different worksheets
    concat_dataframe = pd.concat(sheet_dataframes)
    
    return concat_dataframe
    

def main():
    df = workbook_to_dataframe("../../data/test2.xlsx")
    print(df)


if __name__ == "__main__":
    main()    
